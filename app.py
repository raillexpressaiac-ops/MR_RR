"""
SRPS Cargo - Flask + PostgreSQL Backend
========================================
Combines two systems:
  1. MR Management  (Online GST + Offline MR entries)
  2. RR Manager     (Hamali calculator with per-train bag rate)

Database: PostgreSQL (via psycopg2)
Uses DATABASE_URL environment variable (set automatically by Railway).
"""

import os
import io
import re
import time
import random
import psycopg2
import psycopg2.extras
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file
from flask_cors import CORS
from dotenv import load_dotenv

load_dotenv()

# ----------------------------------------------------------------------------
# App setup
# ----------------------------------------------------------------------------
app = Flask(__name__)
CORS(app)


def get_conn():
    """Open a fresh PostgreSQL connection using DATABASE_URL."""
    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        raise RuntimeError("DATABASE_URL environment variable is not set.")
    conn = psycopg2.connect(database_url)
    return conn


def init_db():
    """
    Apply schema.sql safely.
    Uses CREATE TABLE IF NOT EXISTS and ON CONFLICT DO NOTHING,
    so running this on every startup is harmless and will NOT delete
    existing user data.
    """
    schema_path = os.path.join(os.path.dirname(__file__), "schema.sql")
    if not os.path.exists(schema_path):
        print("[init_db] schema.sql not found, skipping.")
        return

    conn = get_conn()
    try:
        with open(schema_path, "r", encoding="utf-8") as f:
            sql = f.read()

        # Strip all -- comments first, THEN split by semicolon.
        # Without this, chunks that start with a comment block (but contain
        # a CREATE TABLE) get incorrectly filtered out, leaving tables unbuilt.
        sql_clean = re.sub(r'--[^\n]*', '', sql)
        statements = [s.strip() for s in sql_clean.split(";") if s.strip()]

        cur = conn.cursor()
        for stmt in statements:
            cur.execute(stmt)
        conn.commit()
        cur.close()
        print("[init_db] PostgreSQL DB initialized successfully.")
    except Exception as e:
        conn.rollback()
        print(f"[init_db] Schema apply failed: {e}")
    finally:
        conn.close()


# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------
def row_to_dict(row):
    """Turn a psycopg2 RealDictRow into a JSON-friendly dict."""
    if row is None:
        return None
    out = {}
    for k, v in row.items():
        if isinstance(v, datetime):
            out[k] = v.isoformat()
        elif isinstance(v, date):
            out[k] = v.isoformat()
        else:
            out[k] = v
    return out


def get_cursor(conn):
    """Return a RealDictCursor so columns are accessible by name."""
    return conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)


def gen_mr_id():
    """Same shape as JS Date.now().toString() - 13-digit timestamp ms."""
    return str(int(time.time() * 1000))


def gen_rr_id():
    """Mirror JS uid(): 'e_' + base36(time) + 5 random chars"""
    t36 = ""
    n = int(time.time() * 1000)
    while n > 0:
        t36 = "0123456789abcdefghijklmnopqrstuvwxyz"[n % 36] + t36
        n //= 36
    rand = "".join(random.choice("0123456789abcdefghijklmnopqrstuvwxyz") for _ in range(5))
    return f"e_{t36}{rand}"


# ============================================================
# ROUTES — Page Rendering
# ============================================================
@app.route("/")
def home():
    return redirect(url_for("mr_system"))


@app.route("/mr")
def mr_system():
    """SRPS MR Management page (online GST + offline MR)."""
    return render_template("mr_system.html")


@app.route("/rr")
def rr_manager():
    """SRPS RR Manager page (Hamali calculator)."""
    return render_template("rr_manager.html")


# ============================================================
# API — MR SYSTEM: Trains
# ============================================================
@app.route("/api/mr/trains", methods=["GET"])
def mr_trains_list():
    conn = get_conn()
    cur = get_cursor(conn)
    cur.execute("SELECT * FROM mr_trains ORDER BY mode DESC, name ASC")
    rows = cur.fetchall()
    cur.close()
    conn.close()

    trains = []
    for r in rows:
        d = row_to_dict(r)
        if d["mode"] == "online":
            fixed = {
                "taxable":      d["fixed_taxable"],
                "cgst":         d["fixed_cgst"],
                "sgst":         d["fixed_sgst"],
                "igst":         d["fixed_igst"],
                "total_supply": d["fixed_total_supply"],
            }
        else:
            fixed = {
                "weight": d["fixed_weight"],
                "gst":    d["fixed_gst"],
                "mr_amt": d["fixed_mr_amt"],
                "total":  d["fixed_total"],
                "pmode":  d["fixed_pmode"],
            }
        trains.append({
            "no":       d["no"],
            "name":     d["name"],
            "mode":     d["mode"],
            "contract": d["contract"],
            "fixed":    fixed,
        })
    return jsonify(trains)


@app.route("/api/mr/trains", methods=["POST"])
def mr_train_create():
    data = request.get_json(force=True)
    no       = (data.get("no") or "").strip()
    name     = (data.get("name") or "").strip()
    mode     = data.get("mode")
    contract = (data.get("contract") or "").strip()

    if not no or not name or mode not in ("online", "offline"):
        return jsonify({"error": "Train no, name and valid mode are required"}), 400

    conn = get_conn()
    cur = get_cursor(conn)
    try:
        if mode == "online":
            cur.execute("""
                INSERT INTO mr_trains (no, name, mode, contract,
                    fixed_taxable, fixed_cgst, fixed_sgst, fixed_igst, fixed_total_supply)
                VALUES (%s, %s, %s, %s, 0, 0, 0, 0, 0)
            """, (no, name, mode, contract))
        else:
            cur.execute("""
                INSERT INTO mr_trains (no, name, mode, contract,
                    fixed_weight, fixed_gst, fixed_mr_amt, fixed_total, fixed_pmode)
                VALUES (%s, %s, %s, %s, '4 TON', 0, 0, 0, 'CASH')
            """, (no, name, mode, contract))
        conn.commit()
    except psycopg2.IntegrityError:
        conn.rollback()
        return jsonify({"error": f"Train number '{no}' already exists"}), 409
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()
    return jsonify({"status": "ok"}), 201


@app.route("/api/mr/trains/<path:train_no>", methods=["DELETE"])
def mr_train_delete(train_no):
    conn = get_conn()
    cur = get_cursor(conn)
    try:
        cur.execute("DELETE FROM mr_trains WHERE no = %s", (train_no,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        if "foreign key" in str(e).lower():
            return jsonify({"error": "This train has entries; remove them first."}), 409
        raise
    finally:
        cur.close()
        conn.close()
    return jsonify({"status": "ok"})


# ============================================================
# API — MR SYSTEM: Entries
# ============================================================
@app.route("/api/mr/entries", methods=["GET"])
def mr_entries_list():
    conn = get_conn()
    cur = get_cursor(conn)
    cur.execute("SELECT * FROM mr_entries ORDER BY entry_date DESC, created_at DESC")
    rows = cur.fetchall()
    cur.close()
    conn.close()

    out = []
    for r in rows:
        d = row_to_dict(r)
        out.append({
            "id":           d["id"],
            "date":         d["entry_date"],
            "train_no":     d["train_no"],
            "mode":         d["mode"],
            "contract":     d["contract"],
            "space":        d["space"],
            "penalty":      d["penalty"],
            "issue_date":   d["issue_date"],
            # ONLINE
            "invoice":            d["invoice"],
            "service_date":       d["service_date"],
            "taxable":            d["taxable"],
            "cgst":               d["cgst"],
            "sgst":               d["sgst"],
            "igst":               d["igst"],
            "total_value_supply": d["total_value_supply"],
            # OFFLINE
            "side":   d["side"],
            "mr":     d["mr"],
            "weight": d["weight"],
            "gst":    d["gst"],
            "mramt":  d["mramt"],
            "total":  d["total"],
            "pmode":  d["pmode"],
            "remark": d["remark"],
        })
    return jsonify(out)


@app.route("/api/mr/entries", methods=["POST"])
def mr_entry_save():
    """Upsert a single entry. Updates if id exists, else inserts new."""
    e = request.get_json(force=True)

    entry_id = e.get("id") or gen_mr_id()
    mode = e.get("mode")
    if mode not in ("online", "offline"):
        return jsonify({"error": "mode must be 'online' or 'offline'"}), 400

    common = {
        "id":         entry_id,
        "entry_date": e.get("date"),
        "train_no":   e.get("train_no"),
        "mode":       mode,
        "contract":   e.get("contract"),
        "space":      e.get("space"),
        "penalty":    float(e.get("penalty") or 0),
        "issue_date": e.get("issue_date") or None,
    }

    if mode == "online":
        specific = {
            "invoice":            e.get("invoice") or None,
            "service_date":       e.get("service_date") or None,
            "taxable":            float(e.get("taxable") or 0),
            "cgst":               float(e.get("cgst") or 0),
            "sgst":               float(e.get("sgst") or 0),
            "igst":               float(e.get("igst") or 0),
            "total_value_supply": float(e.get("total_value_supply") or 0),
            "side": None, "mr": None, "weight": None, "gst": 0,
            "mramt": 0, "total": 0, "pmode": None, "remark": None,
        }
    else:
        specific = {
            "invoice": None, "service_date": None,
            "taxable": 0, "cgst": 0, "sgst": 0, "igst": 0, "total_value_supply": 0,
            "side":   e.get("side"),
            "mr":     e.get("mr") or None,
            "weight": e.get("weight"),
            "gst":    float(e.get("gst") or 0),
            "mramt":  float(e.get("mramt") or 0),
            "total":  float(e.get("total") or 0),
            "pmode":  e.get("pmode"),
            "remark": e.get("remark"),
        }

    payload = {**common, **specific}

    conn = get_conn()
    cur = get_cursor(conn)
    try:
        cur.execute("""
            INSERT INTO mr_entries (
                id, entry_date, train_no, mode, contract, space, penalty, issue_date,
                invoice, service_date, taxable, cgst, sgst, igst, total_value_supply,
                side, mr, weight, gst, mramt, total, pmode, remark, updated_at
            ) VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP
            )
            ON CONFLICT (id) DO UPDATE SET
                entry_date         = EXCLUDED.entry_date,
                train_no           = EXCLUDED.train_no,
                mode               = EXCLUDED.mode,
                contract           = EXCLUDED.contract,
                space              = EXCLUDED.space,
                penalty            = EXCLUDED.penalty,
                issue_date         = EXCLUDED.issue_date,
                invoice            = EXCLUDED.invoice,
                service_date       = EXCLUDED.service_date,
                taxable            = EXCLUDED.taxable,
                cgst               = EXCLUDED.cgst,
                sgst               = EXCLUDED.sgst,
                igst               = EXCLUDED.igst,
                total_value_supply = EXCLUDED.total_value_supply,
                side               = EXCLUDED.side,
                mr                 = EXCLUDED.mr,
                weight             = EXCLUDED.weight,
                gst                = EXCLUDED.gst,
                mramt              = EXCLUDED.mramt,
                total              = EXCLUDED.total,
                pmode              = EXCLUDED.pmode,
                remark             = EXCLUDED.remark,
                updated_at         = CURRENT_TIMESTAMP
        """, (
            payload["id"], payload["entry_date"], payload["train_no"], payload["mode"],
            payload["contract"], payload["space"], payload["penalty"], payload["issue_date"],
            payload["invoice"], payload["service_date"], payload["taxable"], payload["cgst"],
            payload["sgst"], payload["igst"], payload["total_value_supply"],
            payload["side"], payload["mr"], payload["weight"], payload["gst"],
            payload["mramt"], payload["total"], payload["pmode"], payload["remark"]
        ))
        conn.commit()
    except Exception as ex:
        conn.rollback()
        return jsonify({"error": str(ex)}), 500
    finally:
        cur.close()
        conn.close()

    return jsonify({"status": "ok", "id": entry_id})


@app.route("/api/mr/entries/<entry_id>", methods=["DELETE"])
def mr_entry_delete(entry_id):
    conn = get_conn()
    cur = get_cursor(conn)
    cur.execute("DELETE FROM mr_entries WHERE id = %s", (entry_id,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"status": "ok"})


# ============================================================
# API — RR MANAGER: Trains
# ============================================================
@app.route("/api/rr/trains", methods=["GET"])
def rr_trains_list():
    conn = get_conn()
    cur = get_cursor(conn)
    cur.execute("SELECT id, name, rate FROM rr_trains ORDER BY created_at ASC")
    rows = [row_to_dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify(rows)


@app.route("/api/rr/trains", methods=["POST"])
def rr_train_create():
    data = request.get_json(force=True)
    name = (data.get("name") or "").strip()
    rate = data.get("rate")

    if not name:
        return jsonify({"error": "Station name required"}), 400
    try:
        rate = int(rate)
        if rate <= 0:
            raise ValueError
    except (TypeError, ValueError):
        return jsonify({"error": "Valid rate required"}), 400

    safe = "".join(c if c.isalnum() else "_" for c in name.lower()).strip("_")
    suffix = "".join(random.choice("0123456789abcdefghijklmnopqrstuvwxyz") for _ in range(3))
    new_id = f"{safe}_{suffix}"

    conn = get_conn()
    cur = get_cursor(conn)
    try:
        cur.execute(
            "INSERT INTO rr_trains (id, name, rate) VALUES (%s, %s, %s)",
            (new_id, name, rate)
        )
        conn.commit()
    except psycopg2.IntegrityError:
        conn.rollback()
        return jsonify({"error": "Train already exists"}), 409
    finally:
        cur.close()
        conn.close()

    return jsonify({"id": new_id, "name": name, "rate": rate}), 201


@app.route("/api/rr/trains/<train_id>", methods=["DELETE"])
def rr_train_delete(train_id):
    conn = get_conn()
    cur = get_cursor(conn)
    cur.execute("DELETE FROM rr_trains WHERE id = %s", (train_id,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"status": "ok"})


# ============================================================
# API — RR MANAGER: Entries
# ============================================================
@app.route("/api/rr/entries", methods=["GET"])
def rr_entries_list():
    """Return entries grouped by train_id, matching JS shape: { train_id: [entries...] }."""
    conn = get_conn()
    cur = get_cursor(conn)
    cur.execute("SELECT * FROM rr_entries ORDER BY entry_date DESC, created_at DESC")
    rows = cur.fetchall()
    cur.close()
    conn.close()

    grouped = {}
    for r in rows:
        d = row_to_dict(r)
        train_id = d["train_id"]
        grouped.setdefault(train_id, []).append({
            "id":        d["id"],
            "date":      d["entry_date"],
            "rrNo":      d["rr_no"],
            "from":      d["from_station"],
            "consignor": d["consignor"],
            "to":        d["to_station"],
            "consignee": d["consignee"],
            "bag":       d["bag"],
            "gst":       d["gst"],
            "rrAmt":     d["rr_amt"],
            "weight":    d["weight"],
            "rate":      d["rate"],
            "hamali":    d["hamali"],
            "total":     d["total"],
        })
    return jsonify(grouped)


@app.route("/api/rr/entries", methods=["POST"])
def rr_entry_save():
    e = request.get_json(force=True)
    train_id = e.get("train_id")
    if not train_id:
        return jsonify({"error": "train_id required"}), 400

    rr_no = (e.get("rrNo") or "").strip()
    if not rr_no:
        return jsonify({"error": "RR No is required"}), 400

    bag    = int(e.get("bag") or 0)
    rr_amt = float(e.get("rrAmt") or 0)
    rate   = int(e.get("rate") or 0)
    hamali = bag * rate
    total  = rr_amt + hamali

    entry_id = e.get("id") or gen_rr_id()

    payload = {
        "id":           entry_id,
        "train_id":     train_id,
        "entry_date":   e.get("date") or date.today().isoformat(),
        "rr_no":        rr_no,
        "from_station": (e.get("from") or "").strip() or None,
        "consignor":    (e.get("consignor") or "").strip() or None,
        "to_station":   (e.get("to") or "").strip() or None,
        "consignee":    (e.get("consignee") or "").strip() or None,
        "bag":          bag,
        "gst":          (e.get("gst") or "").strip() or None,
        "rr_amt":       rr_amt,
        "weight":       float(e.get("weight") or 0),
        "rate":         rate,
        "hamali":       hamali,
        "total":        total,
    }

    conn = get_conn()
    cur = get_cursor(conn)
    try:
        cur.execute("""
            INSERT INTO rr_entries (
                id, train_id, entry_date, rr_no, from_station, consignor,
                to_station, consignee, bag, gst, rr_amt, weight, rate, hamali, total, updated_at
            ) VALUES (
                %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, CURRENT_TIMESTAMP
            )
            ON CONFLICT (id) DO UPDATE SET
                train_id     = EXCLUDED.train_id,
                entry_date   = EXCLUDED.entry_date,
                rr_no        = EXCLUDED.rr_no,
                from_station = EXCLUDED.from_station,
                consignor    = EXCLUDED.consignor,
                to_station   = EXCLUDED.to_station,
                consignee    = EXCLUDED.consignee,
                bag          = EXCLUDED.bag,
                gst          = EXCLUDED.gst,
                rr_amt       = EXCLUDED.rr_amt,
                weight       = EXCLUDED.weight,
                rate         = EXCLUDED.rate,
                hamali       = EXCLUDED.hamali,
                total        = EXCLUDED.total,
                updated_at   = CURRENT_TIMESTAMP
        """, (
            payload["id"], payload["train_id"], payload["entry_date"], payload["rr_no"],
            payload["from_station"], payload["consignor"], payload["to_station"],
            payload["consignee"], payload["bag"], payload["gst"],
            payload["rr_amt"], payload["weight"], payload["rate"], payload["hamali"],
            payload["total"]
        ))
        conn.commit()
    except Exception as ex:
        conn.rollback()
        if "foreign key" in str(ex).lower():
            return jsonify({"error": "Invalid train_id"}), 400
        return jsonify({"error": str(ex)}), 500
    finally:
        cur.close()
        conn.close()

    return jsonify({"status": "ok", "id": entry_id})


@app.route("/api/rr/entries/<entry_id>", methods=["DELETE"])
def rr_entry_delete(entry_id):
    conn = get_conn()
    cur = get_cursor(conn)
    cur.execute("DELETE FROM rr_entries WHERE id = %s", (entry_id,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"status": "ok"})


# ============================================================
# Backup — Download all data as Excel
# ============================================================
@app.route("/api/backup")
def backup():
    """Generate and download a full Excel backup of all 4 tables."""
    conn = get_conn()
    cur  = get_cursor(conn)

    # ---- fetch all data ----
    cur.execute("SELECT * FROM mr_trains  ORDER BY mode, name")
    mr_trains = cur.fetchall()

    cur.execute("SELECT * FROM mr_entries ORDER BY entry_date DESC")
    mr_entries = cur.fetchall()

    cur.execute("SELECT * FROM rr_trains  ORDER BY name")
    rr_trains = cur.fetchall()

    cur.execute("SELECT * FROM rr_entries ORDER BY entry_date DESC")
    rr_entries = cur.fetchall()

    cur.close()
    conn.close()

    # ---- build workbook ----
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1B6CA8")
    center_align = Alignment(horizontal="center")

    def add_sheet(wb, title, rows):
        ws = wb.create_sheet(title=title)
        if not rows:
            ws.append(["No data"])
            return
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font  = header_font
            cell.fill  = header_fill
            cell.alignment = center_align
        for row in rows:
            ws.append([str(v) if v is not None else "" for v in row.values()])
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    add_sheet(wb, "MR Trains",   mr_trains)
    add_sheet(wb, "MR Entries",  mr_entries)
    add_sheet(wb, "RR Trains",   rr_trains)
    add_sheet(wb, "RR Entries",  rr_entries)

    # ---- send as download ----
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = date.today().strftime("%Y-%m-%d")
    filename = f"SRPS_Backup_{today}.xlsx"

    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ============================================================
# Health
# ============================================================
@app.route("/api/health")
def health():
    try:
        conn = get_conn()
        cur = get_cursor(conn)
        cur.execute("SELECT 1")
        cur.close()
        conn.close()
        return jsonify({"status": "ok", "db": "connected"})
    except Exception as ex:
        return jsonify({"status": "error", "db": str(ex)}), 500


# ----------------------------------------------------------------------------
# Always run init_db when the module loads so gunicorn (and __main__) both
# get a fully initialised database on startup.
init_db()

if __name__ == "__main__":
    # Local dev only — gunicorn is used in production (see Procfile).
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True, use_reloader=False)
